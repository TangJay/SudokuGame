import numbers
from tabnanny import check
import tkinter as tk
import tkinter.messagebox
from venv import create
import numpy as np
import pandas as pd
import sys
import copy
import time
import random
from solve_sudoku import create_sudoku
from SudokuNumbers import Sudoku
from PIL import ImageTk, Image
import os
from pandas import Index
import time
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from functools import partial

dbPath = os.path.join(os.getcwd(), "SudokuAccounts.xlsx")


class basedesk:
    def __init__(self, master):
        self.root = master
        self.root.title("my window")
        self.root.geometry(
            f"{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}"
        )
        print("x = ", self.root.winfo_screenwidth())
        print("y = ", self.root.winfo_screenheight())
        self.width = self.root.winfo_screenwidth()
        self.height = self.root.winfo_screenheight()
        TitlePage(self.root, self.width, self.height)


class TitlePage:
    def __init__(self, master, width, height):
        self.width = width
        self.height = height
        self.master = master
        self.TitlePage = tk.Frame(
            self.master,
            width=self.master.winfo_screenwidth(),
            height=self.master.winfo_screenheight(),
        )
        self.TitlePage.pack()

        SudokuGame = tk.Label(self.TitlePage, font=("Impact", 70), text="Sudoku Game")
        SudokuGame.place(x=self.width // 3, y=10, anchor="nw")

        account = tk.Label(self.TitlePage, font=("Ink Free", 30), text="account:")
        account.place(x=self.width // 4.7, y=self.height // 3.2, anchor="nw")

        self.account_text = tk.Text(
            self.TitlePage, height=1, font=("Ink Free", 30), width=30
        )
        self.account_text.place(x=self.width // 3.3, y=self.height // 3.15, anchor="nw")

        password = tk.Label(self.TitlePage, font=("Ink Free", 30), text="password:")
        password.place(x=self.width // 5.2, y=self.height // 2.4, anchor="nw")

        self.password_text = tk.Text(
            self.TitlePage, height=1, font=("Ink Free", 30), width=30
        )
        self.password_text.place(x=self.width // 3.3, y=self.height // 2.4, anchor="nw")

        login = tk.Button(
            self.TitlePage,
            width=10,
            height=1,
            text="Login",
            font=("chiller", 50),
            command=self.CheckAccount,
            bg="#F2DF3A",
        ).place(x=self.width // 2.5, y=self.height // 1.7)

        no_account = tk.Button(
            self.TitlePage,
            font=("Ink Free", 30),
            text="no account???    sign up now!",
            bg="#C8FFD4",
            command=self.NoAccount,
        ).place(x=self.width // 3.2, y=self.height // 1.2, anchor="nw")

    def CheckAccount(self):
        AccountResult = self.account_text.get("1.0", "end").strip()
        PasswordResult = self.password_text.get("1.0", "end").strip()
        df = pd.read_excel(dbPath, converters={"account": str, "password": str})
        if AccountResult != None:
            if PasswordResult != None:
                mask1 = df["account"] == f"{AccountResult}"
                mask2 = df["password"] == f"{PasswordResult}"
                # print(mask1, mask2)
                if AccountResult in df["account"].values:
                    id = Index(df["account"]).get_loc(AccountResult)
                    password = df.at[id, "password"]
                    if password == PasswordResult:
                        self.account = AccountResult
                        self.Login()
                else:
                    print(df["account"])
                # tk.messagebox.showinfo(title = "well done!", message = "login complete")
                # self.SignUpAllready()

    def NoAccount(self):
        self.TitlePage.destroy()
        SignUp(self.master)

    def Login(self):
        self.TitlePage.destroy()
        ChooseGameMode(self.master, self.account)


class SignUp:
    def __init__(self, master):
        self.master = master
        self.width = self.master.winfo_screenwidth()
        self.height = self.master.winfo_screenheight()

        self.SignUp = tk.Frame(
            self.master,
            width=self.master.winfo_screenwidth(),
            height=self.master.winfo_screenheight(),
        )
        self.SignUp.pack()

        SignUp = tk.Label(self.SignUp, font=("Impact", 70), text="Sign Up")
        SignUp.place(x=self.width // 2.5, y=10, anchor="nw")

        account = tk.Label(self.SignUp, font=("Ink Free", 30), text="account:")
        account.place(x=self.width // 4.7, y=self.height // 3.2, anchor="nw")

        self.account_text = tk.Text(
            self.SignUp, height=1, font=("Ink Free", 30), width=30
        )
        self.account_text.place(x=self.width // 3.3, y=self.height // 3.15, anchor="nw")

        password = tk.Label(self.SignUp, font=("Ink Free", 30), text="password:")
        password.place(x=self.width // 5.2, y=self.height // 2.4, anchor="nw")

        self.password_text = tk.Text(
            self.SignUp, height=1, font=("Ink Free", 30), width=30
        )
        self.password_text.place(x=self.width // 3.3, y=self.height // 2.4, anchor="nw")

        check = tk.Label(
            self.SignUp, font="Castellar, 30", text="please check your password again"
        )
        check.place(x=self.width // 3, y=self.height // 2 + 10)

        password2 = tk.Label(self.SignUp, font=("Ink Free", 30), text="password:")
        password2.place(x=self.width // 5.2, y=self.height // 1.6, anchor="nw")

        self.password2_text = tk.Text(
            self.SignUp, height=1, font=("Ink Free", 30), width=30
        )
        self.password2_text.place(
            x=self.width // 3.3, y=self.height // 1.6, anchor="nw"
        )

        confirm = tk.Button(
            self.SignUp,
            width=10,
            height=1,
            text="SignUp",
            fg="RoyalBlue",
            font=("chiller", 50),
            command=self.CheckAccount,
            bg="gold",
        )
        confirm.place(x=self.width // 2.5, y=self.height // 1.3)

    def CheckAccount(self):
        AccountResult = self.account_text.get("1.0", "end").strip()
        Password1Result = self.password_text.get("1.0", "end").strip()
        Password2Result = self.password2_text.get("1.0", "end").strip()
        df = pd.read_excel(dbPath, converters={"account": str, "password": str})
        PasswordCharacter = list()
        lower = 0
        higher = 0
        integer = 0
        if AccountResult == "":
            PasswordCharacter.append("no account")
        if Password1Result == "":
            PasswordCharacter.append("no password1")
        if Password2Result == "":
            PasswordCharacter.append("no password2")
        if Password1Result != Password2Result:
            PasswordCharacter.append("passwords are not the same")
        print(type(df["account"].values))
        if AccountResult.lower() in [name.lower() for name in df["account"].values]:
            PasswordCharacter.append("account already exists")
        for strs in Password1Result:
            if strs.islower():
                lower += 1
            if strs.isupper():
                higher += 1
            if strs.isdigit():
                integer += 1
        if lower <= 0:
            PasswordCharacter.append("no LowerAlphabet")
        if higher <= 0:
            PasswordCharacter.append("no UpperAlphabet")
        if integer <= 0:
            PasswordCharacter.append("no integer")
        if len(Password1Result) < 8:
            PasswordCharacter.append("password too short")

        if len(PasswordCharacter) == 0:
            new_df = df.append(
                {
                    "account": AccountResult,
                    "password": Password1Result,
                    "save1": "False",
                    "save2": "False",
                    "save3": "False",
                    "save4": "False",
                    "save5": "False",
                },
                ignore_index=True,
            )
            AccountData = {
                "saveID": [],
                "time": [],
                "TimeUsed": [],
                "CompletedPercentage": [],
            }
            for i in range(81):
                AccountData[f"value{i // 9}-{i % 9}"] = 0
            AccountDf = pd.DataFrame(AccountData)
            wb = openpyxl.load_workbook(dbPath)
            removeSheet = wb["account"]
            wb.remove(removeSheet)
            ws = wb.create_sheet("account")

            for row in dataframe_to_rows(new_df, index=False):
                ws.append(row)
            # writer = pd.ExcelWriter(dbPath)
            # new_df.to_excel(writer, index=False, sheet_name = "account")
            # new_df.to_excel(writer, index=False, sheet_name = "account1")
            # writer.close()
            if AccountResult not in wb.sheetnames:
                ws = wb.create_sheet(AccountResult)
                for row in dataframe_to_rows(AccountDf, index=False):
                    ws.append(row)
            wb.save(dbPath)
            wb.close()
            print(1)
            # sys.exit()

            # df = pd.DataFrame(data)
            # df.to_excel(dbPath, index = False, sheet_name = AccountResult)
            self.account = AccountResult
            self.SignUpAllready()
        else:
            ss = ""
            for res in PasswordCharacter:
                ss += f"{res}\n"
            tk.messagebox.showerror(title="error", message=f"{ss}")

    def SignUpAllready(self):
        self.SignUp.destroy()
        ChooseGameMode(self.master, self.account)


class ChooseGameMode:
    def __init__(self, master, account):
        self.GameDifficulty = None
        self.account = account
        self.master = master

        self.width = self.master.winfo_screenwidth()
        self.height = self.master.winfo_screenheight()

        self.buttonsShowed = False
        self.ChooseGameMode = tk.Frame(
            self.master,
            width=self.master.winfo_screenwidth(),
            height=self.master.winfo_screenheight(),
        )
        self.ChooseGameMode.pack()

        ChooseGameMode = tk.Label(
            self.ChooseGameMode, font=("Impact", 70), text="SudokuGame"
        )
        ChooseGameMode.place(x=600, y=10, anchor="nw")

        self.NewGame = tk.Button(
            self.ChooseGameMode,
            width=10,
            height=2,
            text="New Game",
            bg="#B1AFFF",
            activebackground="#B1AFFF",
            font=("algerian, 50"),
            command=self.newgame,
        )
        self.NewGame.place(x=self.width // 3.9, y=self.height // 3.7, anchor="nw")
        self.LoadGame = tk.Button(
            self.ChooseGameMode,
            width=10,
            height=2,
            text="Load Game",
            bg="#B1AFFF",
            activebackground="#B1AFFF",
            font=("algerian, 50"),
            command=self.loadgame,
        )
        self.LoadGame.place(x=self.width // 1.9, y=self.height // 3.7, anchor="nw")

    def newgame(self):
        self.SaveFile = None
        self.NewGame.configure(state="disabled")
        self.LoadGame.configure(state="disabled")
        self.CurrentHighlightObject = None
        self.diffculties = list()
        self.SaveFrame = tk.Frame(
            self.ChooseGameMode, height=self.height, width=self.width / 3, bg="#F2DF3A"
        )
        self.SaveFrame.place(x=self.width / 3, y=0)
        self.easy = tk.Button(
            self.SaveFrame,
            width=8,
            height=1,
            text="easy",
            bg="#C8FFD4",
            font=("algerian, 70"),
            command=partial(self.HighlightDifficulty, 0, "easy"),
        )
        self.easy.place(x=self.width / 22, y=self.height / 5 * 1)
        self.diffculties.append(self.easy)
        self.normal = tk.Button(
            self.SaveFrame,
            width=8,
            height=1,
            text="normal",
            bg="#C8FFD4",
            font=("algerian, 70"),
            command=partial(self.HighlightDifficulty, 1, "normal"),
        )
        self.normal.place(x=self.width / 22, y=self.height / 5 * 2)
        self.diffculties.append(self.normal)
        self.hard = tk.Button(
            self.SaveFrame,
            width=8,
            height=1,
            text="hard",
            bg="#C8FFD4",
            font=("algerian, 70"),
            command=partial(self.HighlightDifficulty, 2, "hard"),
        )
        self.hard.place(x=self.width / 22, y=self.height / 5 * 3)
        self.diffculties.append(self.hard)
        back_to_game = tk.Button(
            self.SaveFrame,
            font="Verdana, 30",
            bg="#0078AA",
            command=self.BackToGame,
            height=1,
            width=10,
            text="back",
        )
        back_to_game.place(x=self.width / 43, y=self.height / 1.225)
        confirm = tk.Button(
            self.SaveFrame,
            font="Verdana, 30",
            bg="#0078AA",
            command=self.SudokuPage,
            height=1,
            width=10,
            text="confirm",
        )
        confirm.place(x=self.width / 5.5, y=self.height / 1.225)

    def HighlightDifficulty(self, choose, difficulty):
        self.GameDifficulty = difficulty
        self.NewHighlightObject = self.diffculties[choose]
        if self.CurrentHighlightObject == None:
            self.NewHighlightObject.configure(bg="#B1AFFF")
            print(1)
        elif self.NewHighlightObject != self.CurrentHighlightObject:
            self.CurrentHighlightObject.configure(bg="#9DF1DF")
            self.NewHighlightObject.configure(bg="#B1AFFF")
            print(2)
        self.CurrentHighlightObject = self.NewHighlightObject

    def Normal(self):
        self.SudokuPage("normal")

    def Hard(self):
        self.SudokuPage("hard")

    def loadgame(self):
        self.NewGame.configure(state="disabled")
        self.LoadGame.configure(state="disabled")
        self.CurrentHighlightObject = None
        self.SaveFileFrameList = list()
        self.SaveFrame = tk.Frame(
            self.ChooseGameMode, height=self.height, width=self.width / 3, bg="#F2DF3A"
        )
        self.SaveFrame.place(x=self.width / 3, y=0)
        df = pd.read_excel(
            dbPath,
            converters={"account": str, "password": str},
            sheet_name=self.account,
        )
        print(df[["saveID"]])
        self.SaveList = list()
        self.SaveIndexList = list()
        for i in range(1, 6):
            if i not in (df[["saveID"]].values):
                print("AAA")
                self.SaveIndexList.append("empty")
                SaveIndex = "empty"
            else:
                print("BBB")
                saveIDDataframe = df.loc[df["saveID"] == i]
                print("saveIDDataframe = ", saveIDDataframe)
                print("type of saveIDDataframe = ", type(saveIDDataframe))
                LastPlayedTime = saveIDDataframe[["time"]].values[0]
                percentage = (
                    f"completeness: {saveIDDataframe['CompletedPercentage'].values[0]}%"
                )
                self.SaveIndexList.append(f"{LastPlayedTime}\n{percentage}")
                SaveIndex = f"{LastPlayedTime}\n{percentage}"
                print(SaveIndex)
            SaveFileFrame = tk.Button(
                self.SaveFrame,
                font="Verdana, 15",
                command=lambda c=i: self.highlight(c),
                bg="#9DF1DF",
                height=5,
                width=37,
                text=f"save{i}\n\n{SaveIndex}",
            )
            self.SaveFileFrameList.append(SaveFileFrame)
            SaveFileFrame.place(x=self.width / 22, y=self.height / 7 * (i - 0.5))
        back_to_game = tk.Button(
            self.SaveFrame,
            font="Verdana, 30",
            bg="#0078AA",
            command=self.BackToGame,
            height=1,
            width=10,
            text="back",
        )
        back_to_game.place(x=self.width / 43, y=self.height / 1.225)
        confirm = tk.Button(
            self.SaveFrame,
            font="Verdana, 30",
            bg="#0078AA",
            command=self.LoadingGame,
            height=1,
            width=10,
            text="confirm",
        )
        confirm.place(x=self.width / 5.5, y=self.height / 1.225)

    def BackToGame(self):
        self.CurrentHighlightObject = None
        self.GameDifficulty = None
        self.SaveFile = None
        self.NewGame.configure(state="active")
        self.LoadGame.configure(state="active")
        self.SaveFrame.destroy()

    def LoadingGame(self):
        if self.SaveFile == None:
            tk.messagebox.showerror(title="error", message="please choose a file first")
            return
        self.ChooseGameMode.destroy()
        GamePage(self.master, None, self.account, "LoadGame", self.SaveFile)

    def highlight(self, choose):
        self.NewHighlightObject = self.SaveFileFrameList[choose - 1]
        if self.CurrentHighlightObject == None:
            self.NewHighlightObject.configure(bg="#B1AFFF")
        elif self.NewHighlightObject != self.CurrentHighlightObject:
            self.CurrentHighlightObject.configure(bg="#9DF1DF")
            self.NewHighlightObject.configure(bg="#B1AFFF")
        self.CurrentHighlightObject = self.NewHighlightObject
        self.SaveFile = choose

    def SudokuPage(self):
        if self.GameDifficulty == None:
            tk.messagebox.showerror(
                title="error", message="please choose a difficulty first"
            )
            return
        self.ChooseGameMode.destroy()
        GamePage(self.master, self.GameDifficulty, self.account, "NewGame", None)


class GamePage:
    def __init__(self, master, difficulty, account, GameType, SaveNumber):
        self.step = list()
        self.master = master
        self.account = account
        self.width = self.master.winfo_screenwidth()
        self.height = self.master.winfo_screenheight()
        self.difficulty = difficulty
        self.GameType = GameType
        self.SaveNumber = SaveNumber

        self.Game = tk.Frame(
            self.master,
            width=self.master.winfo_screenwidth(),
            height=self.master.winfo_screenheight(),
        )
        self.Game.pack()

        if self.GameType == "NewGame":
            self.NewGame()
        elif self.GameType == "LoadGame":
            self.LoadGame()

    def LoadGame(self):
        self.sudoku = Sudoku()
        df = pd.read_excel(dbPath, sheet_name=self.account)
        self.nums = df.loc[df["nums"] == self.SaveNumber]
        self.saveIDDataframe = df.loc[df["saveID"] == self.SaveNumber]
        while not self.sudoku.make_digits():
            pass
        for i in range(9):
            for j in range(9):
                print(type(self.saveIDDataframe))
                print(self.saveIDDataframe.at[0, f"value{i}-{j}"])
                value = str(self.saveIDDataframe[f"value{i}-{j}"]).split("/")[0]
                print(f"value = {value}")
                print(f"value.type = {type(value)}")
                self.sudoku.digits[i][j] = int(value)
        self.MakeBlock()

    def NewGame(self):
        if self.difficulty == "hard":
            self.nums = random.randint(16, 20)
        elif self.difficulty == "normal":
            self.nums = random.randint(21, 25)
        elif self.difficulty == "easy":
            self.nums = random.randint(26, 30)
        self.sudoku = Sudoku()
        while not self.sudoku.make_digits():
            pass
        zeros = 0
        while zeros < 81 - self.nums:
            x = random.randint(0, 8)
            y = random.randint(0, 8)
            if self.sudoku.digits[x][y] != 0:
                zeros += 1
            self.sudoku.digits[x][y] = 0
        self.MakeBlock()
        data = {
            "0": [],
            "1": [],
            "2": [],
            "3": [],
            "4": [],
            "5": [],
            "6": [],
            "7": [],
            "8": [],
        }
        for i in range(9):
            for j in range(9):
                data[f"{i}"].append(self.sudoku.digits[i][j])
        df = pd.DataFrame(data)

    def places(self, row, column):
        self.row = row
        self.column = column

    def MakeBlock(self):
        self.ButtonList = list()
        self.StartTime = time.time()
        bigFrame = tk.Frame(self.Game, width=1200, height=1000, bd=10, bg="#FFEA20")
        bigFrame.propagate(0)
        bigFrame.place(x=350, y=10)
        self.smallFrames = dict()
        self.smallFrames_number = dict()
        for i in range(9):
            row = i // 3 + 1
            column = i % 3 + 1
            mediumFrame = tk.Frame(bigFrame, width=270, height=270, bd=8, bg="#0078AA")
            mediumFrame.propagate(0)
            mediumFrame.grid(row=row, column=column, padx=5, pady=5)
            for j in range(9):
                smallrow = j // 3 + 1
                smallcolumn = j % 3 + 1
                Row = (row - 1) * 3 + (smallrow - 1)
                Column = (column - 1) * 3 + (smallcolumn - 1)
                if self.sudoku.digits[Row][Column] == 0:
                    show = " "
                else:
                    show = self.sudoku.digits[Row][Column]
                text = tk.StringVar()
                text.set(show)
                self.smallFrames_number[(Row, Column)] = text
                sf = tk.Button(
                    mediumFrame,
                    width=3,
                    height=1,
                    bd=5,
                    bg="#9DF1DF",
                    font="Ariel, 32",
                    textvariable=text,
                    command=lambda row=Row, column=Column: self.places(row, column),
                )
                self.smallFrames[(Row, Column)] = sf
                sf.propagate(0)
                sf.grid(row=smallrow, column=smallcolumn, padx=5, pady=5)
                if self.GameType == "LoadGame":
                    disabled = self.saveIDDataframe[f"value{i}-{j}"].split("/")[1]
                    if disabled == "T":
                        sf.configure(state="disabled")
                if self.GameType == "NewGame":
                    if text.get() != " ":
                        sf.configure(state="disabled")
                self.ButtonList.append(sf)

        mediumFrame = tk.Frame(self.Game, width=270, height=270, bd=8, bg="#0078AA")
        mediumFrame.propagate(0)
        mediumFrame.place(x=1350, y=600)
        for j in range(9):
            smallrow = j // 3
            smallcolumn = j % 3
            if (
                self.sudoku.digits[(row - 1) * 3 + (smallrow - 1)][
                    (column - 1) * 3 + (smallcolumn - 1)
                ]
                == 0
            ):
                show = " "
            else:
                show = self.sudoku.digits[(row - 1) * 3 + (smallrow - 1)][
                    (column - 1) * 3 + (smallcolumn - 1)
                ]
            smallFrame = tk.Button(
                mediumFrame,
                width=3,
                height=1,
                bd=5,
                bg="#9DF1DF",
                text=j + 1,
                font="Ariel, 32",
                command=lambda numbers=j + 1: self.numbers(numbers),
            )
            smallFrame.propagate(0)
            smallFrame.grid(row=smallrow, column=smallcolumn, padx=5, pady=5)
        """
        image = Image.open("C:/Users/Jay/OneDrive/桌面/python_class/img/tomato.jpg")
        size_thumbnail = (300, 300)
        image.thumbnail(size_thumbnail)
        test = ImageTk.PhotoImage(image)
        
        img= (Image.open("C:/Users/Jay/OneDrive/桌面/python_class/img/tomato.jpg"))
        resized_image= img.resize((300,205), Image.ANTIALIAS)
        image= ImageTk.PhotoImage(resized_image)
        """
        # tk.Label(self.Game, image = image, height = 10, width = 10).place(x = 30 , y = 30)
        backButton = tk.Button(
            self.Game, text="back", font="chiller, 30", bg="#3AB4F2", command=self.back
        ).place(x=30, y=30)

        self.SaveButton = tk.Button(
            self.Game, text="save", bg="#3AB4F2", font="chiller, 30", command=self.save
        ).place(x=1570, y=30)

    def save(self):
        self.SaveFrame = tk.Frame(
            self.Game, height=self.height, width=self.width / 3, bg="#F2DF3A"
        )
        self.SaveFrame.place(x=self.width / 3, y=0)
        df = pd.read_excel(
            dbPath,
            converters={"account": str, "password": str},
            sheet_name=self.account,
        )
        print(df[["saveID"]])
        self.SaveList = list()
        self.SaveIndexList = list()
        for i in range(1, 6):
            if i not in (df[["saveID"]].values):
                print("AAA")
                self.SaveIndexList.append("empty")
                SaveIndex = "empty"
            else:
                print("BBB")
                saveIDDataframe = df.loc[df["saveID"] == i]
                print("saveIDDataframe = ", saveIDDataframe)
                print("type of saveIDDataframe = ", type(saveIDDataframe))
                LastPlayedTime = saveIDDataframe[["time"]].values[0]
                percentage = (
                    f"completeness: {saveIDDataframe['CompletedPercentage'].values[0]}%"
                )
                self.SaveIndexList.append(f"{LastPlayedTime}\n{percentage}")
                SaveIndex = f"{LastPlayedTime}\n{percentage}"
                print(SaveIndex)
            SaveFileFrame = tk.Button(
                self.SaveFrame,
                font="Verdana, 15",
                command=lambda c=i: self.saving(c),
                bg="#9DF1DF",
                height=5,
                width=37,
                text=f"save{i}\n\n{SaveIndex}",
            )
            SaveFileFrame.place(x=self.width / 22, y=self.height / 7 * (i - 0.5))
            self.SaveList.append(SaveFileFrame)
        back_to_game = tk.Button(
            self.SaveFrame,
            font="Verdana, 30",
            bg="#0078AA",
            command=self.BackToGame,
            height=1,
            width=10,
            text="back to game",
        )
        back_to_game.place(x=self.width / 43, y=self.height / 1.225)
        close_game = tk.Button(
            self.SaveFrame,
            font="Verdana, 30",
            bg="#0078AA",
            command=self.CloseGame,
            height=1,
            width=10,
            text="close game",
        )
        close_game.place(x=self.width / 5.5, y=self.height / 1.225)

    def BackToGame(self):
        self.SaveFrame.destroy()

    def CloseGame(self):
        sys.exit()

    # def saving(self, number):
    #     zeros = 0
    #     self.Endtime = time.time()
    #     self.UsedTime = int(self.StartTime - self.Endtime)
    #     dfdata = pd.read_excel(dbPath, converters={"account":str,"password":str})
    #     wb = openpyxl.load_workbook(dbPath)
    #     print(type(wb.sheetnames))
    #     with pd.ExcelWriter(engine='openpyxl', path = dbPath, mode='w') as writer:

    #         if "jay" not in wb.sheetnames:
    #             data = {"saveID": [], "time": [], "TimeUsed": [], "CompletedPercentage": []}
    #             for i in range(81):
    #                 data[f"value{i // 9}-{i % 9}"] = []
    #                 if self.sudoku.digits[i // 9][i % 9] == 0:
    #                     zeros += 1
    #             df = pd.DataFrame(data)
    #             df.to_excel(writer, index = False, sheet_name = "jay")
    #         dfdata = pd.read_excel(dbPath, sheet_name = "jay")
    #         print("start remove")
    #         wb.remove(wb["jay"])
    #         wb.save(dbPath)
    #         print("end remove")
    #         sys.exit()
    #         seconds = time.time()
    #         local_time = time.ctime(seconds)
    #         self.percentage = int(100 / ((81 - self.nums) * ((81 - self.nums) - zeros)))
    #         #df.to_excel(writer, index=False, sheet_name = "jay")
    #         print("dfdata's type = ",type(dfdata))
    #         new_df = dfdata.append({"saveID": number, "time": local_time, "TimeUsed": self.UsedTime, "CompletedPercentage": self.percentage}, ignore_index=True)
    #         for i in range(81):
    #             new_df[f"value{i // 9}-{i % 9}"] = self.sudoku.digits[i // 9][i % 9]
    #         print(new_df)
    #         new_df.to_excel(writer, index=False, sheet_name = "jay")
    def saving(self, number):
        df = pd.read_excel(
            dbPath,
            converters={"account": str, "password": str},
            sheet_name=self.account,
        )
        print("number = ", number)
        # get time result
        localTime = time.ctime(time.time())
        self.EndTime = time.time()
        usedTime = int(self.EndTime - self.StartTime)

        # calculate complete percentage
        zeros = 0
        for i in range(81):
            if self.sudoku.digits[i // 9][i % 9] == 0:
                zeros += 1
        self.percentage = ((81 - zeros - self.nums) / (81 - self.nums)) * 100

        # generate new row
        newRow = {
            "saveID": number,
            "time": localTime,
            "TimeUsed": usedTime,
            "CompletedPercentage": self.percentage,
            "nums": self.nums,
        }
        for i in range(9):
            for j in range(9):
                if self.ButtonList[i * 9 + j]["state"] == "disabled":
                    disabled = "T"
                else:
                    disabled = "F"
                newRow[f"value{i}-{j}"] = f"{self.sudoku.digits[j][i]}/{disabled}"
        print(newRow)
        new_df = df.append(newRow, ignore_index=True)
        print("number = ", number)

        # start to save file
        wb = openpyxl.load_workbook(dbPath)
        removeSheet = wb[str(self.account)]
        wb.remove(removeSheet)
        ws = wb.create_sheet(str(self.account))

        for row in dataframe_to_rows(new_df, index=False):
            ws.append(row)
        """
        wb = openpyxl.load_workbook(dbPath)
        print('create workbook')
        # writer = pd.ExcelWriter(dbPath, engine = 'openpyxl')
        # writer.book = wb
        print('sheetnames:', wb.sheetnames)
        df = None
        """
        if self.account in wb.sheetnames:
            print("account in sheets")
            df = pd.read_excel(dbPath, sheet_name=self.account)
            wb.remove(wb[self.account])
            print("ori df")
            print(df)
            df = df[df.saveID != number]
            df = df.append(newRow, ignore_index=True)
            print("df with new index")
            print(df)
            ws = wb.create_sheet(self.account)
            for row in dataframe_to_rows(df, index=False):
                ws.append(row)
            wb.save(dbPath)
            wb.close()

        else:
            print("account not in sheet")
            columns = ["saveID", "time", "TimeUsed", "CompletedPercentage"]
            for i in range(9):
                for j in range(9):
                    columns.append(f"value{i}-{j}")
            print(f"columns = {columns}")
            df = pd.DataFrame(columns=columns)
            df = df.append(newRow, ignore_index=True)
            print("new df")
            print(df)
            ws = wb.create_sheet(self.account)
            for row in dataframe_to_rows(df, index=False):
                ws.append(row)
            wb.save(dbPath)  # save workbook
            wb.close()  # close workbook
        if number not in (df[["saveID"]].values):
            SaveIndex = "empty"
        else:
            saveIDDataframe = df.loc[df["saveID"] == number]
            LastPlayedTime = saveIDDataframe[["time"]].values[0]
            percentage = (
                f"completeness: {saveIDDataframe['CompletedPercentage'].values[0]}%"
            )
            SaveIndex = f"{LastPlayedTime}\n{percentage}"

        print(f"self.savelist = {self.SaveList}")
        self.SaveList[number - 1].destroy()
        SaveFileFrame = tk.Button(
            self.SaveFrame,
            font="Verdana, 15",
            command=lambda c=number: self.saving(c),
            bg="#9DF1DF",
            height=5,
            width=37,
            text=f"save{number}\n\n{SaveIndex}",
        )
        SaveFileFrame.place(x=self.width / 22, y=self.height / 7 * (number - 0.5))
        print("current = ", self.SaveList)

    def back(self):
        lastStep = self.step.pop()
        if lastStep[2] == 0:
            show = " "
        else:
            show = lastStep[2]
        self.sudoku.digits[lastStep[0]][lastStep[1]] = lastStep[2]
        self.smallFrames_number[lastStep[0], lastStep[1]].set(show)

    def numbers(self, numbers):
        if self.row == None or self.column == None:
            tk.messagebox.showerror(title="error", message="please choose a box first")
            return
        numbers_can = self.check_clues()
        if not numbers in numbers_can:
            tk.messagebox.showerror(
                title="error", message="you cannot use that number in the block"
            )
            return
        self.step.append(
            [self.row, self.column, self.sudoku.digits[self.row][self.column]]
        )
        self.sudoku.digits[self.row][self.column] = numbers
        self.smallFrames_number[(self.row, self.column)].set(numbers)

        for i in range(9):
            for j in range(9):
                if self.sudoku.digits[i][j] == 0:
                    return
        tk.messagebox.showinfo(title="Victory!", message="you have solved the sudoku!")
        self.Game.destroy()
        ChooseGameMode(self.master)

    def check_clues(self):
        clues = set()
        num = [1, 2, 3, 4, 5, 6, 7, 8, 9]
        for i in range(9):
            if self.sudoku.digits[i][self.column] != 0:
                clues.add(self.sudoku.digits[i][self.column])
        for i in range(9):
            if self.sudoku.digits[self.row][i] != 0:
                clues.add(self.sudoku.digits[self.row][i])
        b_x = self.row // 3
        b_y = self.column // 3
        for i in range(3):
            for j in range(3):
                if self.sudoku.digits[b_x * 3 + j][b_y * 3 + i] != 0:
                    clues.add(self.sudoku.digits[b_x * 3 + j][b_y * 3 + i])
        return set(num) - set(clues)

        """
        blocks = [[None for x in range(9)] for x in range(9)] 
        btn_cells = [[None for x in range(9)] for x in range(9)]        
        for i in range(9): # 0~8
            for j in range(9): # 0~8
                var = tk.IntVar()
                blocks[i][j] = tk.Frame(self.master,width = 110, height = 110, bd=3, highlightbackground='light blue',
                                        highlightcolor='light blue', highlightthickness=1, bg = "lime")
                blocks[i][j].propagate(0)
                blocks[i][j].place(x = 300 + i * 100, y = j * 100)
                
                btn_cells = tk.Button(blocks[i][j], bg='white', textvariable=var, width = 17, height = 6)
                btn_cells.pack(side='bottom', fill='both')
                if self.sudoku.digits[i][j] == 0:
                    show = " "
                else:
                    show = self.sudoku.digits[i][j]
                var.set(show)
                
        # Add the 9 * 9 cellss
        
        for j in range(9):
            for i in range(9):
                # Add cell to the block
                # Add a frame so that the cell can form a square
                
                frm_cell = tk.Frame(blocks[i // 1][j // 9])
                frm_cell.grid(row=(i % 9), column=(j % 1), sticky='nsew')
                frm_cell.rowconfigure(0, minsize=96, weight=1)
                frm_cell.columnconfigure(0, minsize=96, weight=1)
                """


if __name__ == "__main__":
    if os.path.exists(dbPath) == False:
        data = {
            "account": [],
            "password": [],
            "save1": [],
            "save2": [],
            "save3": [],
            "save4": [],
            "save5": [],
        }
        df = pd.DataFrame(data)
        print(df)
        df.to_excel(dbPath, index=False, sheet_name="account")
    else:
        df = pd.read_excel(dbPath)
        print(df)
    print(dbPath)
    sys.setrecursionlimit(200000)
    root = tk.Tk()
    basedesk(root)
    root.mainloop()

    # NewGame 換成視窗
    # LoadingGame 偵測選擇
    # 兩個視窗不重複
    #
